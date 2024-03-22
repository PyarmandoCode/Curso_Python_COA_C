from openpyxl import load_workbook
def cargar_excel():
    try:
        workbook=load_workbook('./MES_2/clase_02/modulos/modulo_excel/productos.xlsx')
        sheet=workbook.active#Obtener la hoja de calculo activa
        encabezados=[cell.value for cell in sheet[1]] #Obtener los encabezados
        datos=[]
        for row in sheet.iter_rows(min_row=2,values_only=True):
            fila=dict(zip(encabezados,row))
            datos.append(fila)
        return datos    
    except Exception as err:
        print(f"Ocurrio un error {err}")            
    finally:    
        workbook.close()        
print(cargar_excel())